from django.shortcuts import render, redirect , HttpResponse 
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook 
from .models import Order
import os
import openpyxl

def upload_file(request):
    if request.method == 'POST' and request.FILES['file']:
        uploaded_file = request.FILES['file']
        insert_data_from_excel(uploaded_file)

    orders = Order.objects.all()
    return render(request, 'index.html', {'orders': orders})

def insert_data_from_excel(uploaded_file):
    wb = load_workbook(uploaded_file)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        order_id, product_name, product_price, shipped = row

        
        Order.objects.create(
            order_id=order_id,
            product_name=product_name,
            product_price=product_price,
            shipped=shipped
        )      

def download_file(request):
    
    response = create_excel_file()
    response['Content-Disposition'] = 'attachment; filename=orders.xlsx'
    return response

def create_excel_file():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(['Order ID', 'Product Name', 'Product Price', 'Shipped'])

    orders = Order.objects.all()

    for order in orders:
        sheet.append([order.order_id, order.product_name, order.product_price, order.shipped])

    file_path = 'media/orders.xlsx'
    wb.save(file_path)

    with open(file_path, 'rb') as file:
        response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    os.remove(file_path)
    return response